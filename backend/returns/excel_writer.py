"""Write a filing-ready, multi-sheet Excel workbook from the analysis result."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_TITLE_FONT = Font(bold=True, size=13, color="1F4E78")


def _style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws) -> None:
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(width + 3, 45)


def _table(ws, title: str, headers: list[str], rows: list[list], start: int) -> int:
    ws.cell(row=start, column=1, value=title).font = _TITLE_FONT
    hr = start + 1
    for i, h in enumerate(headers, 1):
        ws.cell(row=hr, column=i, value=h)
    _style_header(ws, hr, len(headers))
    r = hr + 1
    for row in rows:
        for i, v in enumerate(row, 1):
            ws.cell(row=r, column=i, value=v)
        r += 1
    return r + 2  # blank spacer


def write_filing_excel(path: str | Path, analysis: dict) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # ── GSTR-3B summary ──
    ws = wb.active
    ws.title = "GSTR-3B Summary"
    d = analysis["gstr3b_detail"]
    liab = analysis["gstr3b"]
    r = _table(ws, "GSTR-3B — Summary of Tax Liability", ["Particulars", "Amount (₹)"], [
        ["Outward taxable value", liab["outward_taxable_value"]],
        ["Outward tax (output)", liab["outward_tax"]],
        ["Gross ITC in books", d["gross_itc_in_books"]],
        ["Less: ITC blocked u/s 17(5)", d["itc_blocked_17_5"]],
        ["Less: ITC at risk (2A mismatch/missing)", d["itc_at_risk_2a_mismatch"]],
        ["Net ITC claimable", d["net_itc_claimable"]],
        ["NET TAX PAYABLE", d["net_tax_payable"]],
    ], 1)
    _autosize(ws)

    # ── GSTR-1: B2B ──
    ws = wb.create_sheet("GSTR-1 B2B")
    g1 = analysis["gstr1"]
    _table(ws, "GSTR-1 — B2B Invoices", [
        "GSTIN", "Recipient", "Invoice No", "Date", "Taxable", "CGST", "SGST", "IGST", "Rate %",
    ], [
        [b["gstin"], b["recipient"], b["invoice_number"], b["invoice_date"],
         b["taxable_value"], b["cgst"], b["sgst"], b["igst"], b["rate"]]
        for b in g1["b2b_invoices"]
    ], 1)
    _autosize(ws)

    # ── GSTR-1: B2C + rate-wise + HSN ──
    ws = wb.create_sheet("GSTR-1 B2C & HSN")
    r = _table(ws, "B2C — Rate-wise", ["Rate %", "Taxable", "Tax", "Count"],
               [[k, v["taxable_value"], v["tax"], v["count"]] for k, v in sorted(g1["b2c_rate_wise"].items())], 1)
    r = _table(ws, "HSN Summary", ["HSN/SAC", "Taxable", "Tax", "Count"],
               [[k, v["taxable_value"], v["tax"], v["count"]] for k, v in g1["hsn_summary"].items()], r)
    _autosize(ws)

    # ── Reconciliation ──
    ws = wb.create_sheet("Reconciliation")
    _table(ws, "Purchase vs GSTR-2A Reconciliation", [
        "Invoice", "GSTIN", "Supplier", "Status", "Book ITC", "GSTR-2A ITC", "Diff", "At Risk", "Note",
    ], [
        [l["invoice_number"], l["gstin"], l["supplier"], l["status"],
         l["book_tax"], l["gstr2a_tax"], l["tax_difference"], l["itc_at_risk"], l["note"]]
        for l in analysis["reconciliation"]["lines"]
    ], 1)
    _autosize(ws)

    # ── Audit log (cited decisions) ──
    ws = wb.create_sheet("Audit Log")
    _table(ws, "Decision Audit Trail (with GST Act citations)", [
        "Agent", "Invoice", "Decision", "Rationale", "Cited Section", "Citation Snippet", "Confident",
    ], [
        [e["agent"], e["entity_ref"], e["decision"], e["rationale"],
         (e["citation"]["section"] if e.get("citation") else ""),
         (e["citation"]["snippet"] if e.get("citation") else ""),
         "Yes" if e["confident"] else "No (review)"]
        for e in analysis["audit_log"]
    ], 1)
    _autosize(ws)

    wb.save(path)
    return path
